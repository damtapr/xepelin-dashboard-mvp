from io import BytesIO
from pathlib import Path
import runpy
import pandas as pd
import streamlit as st

# Paths esperados por tu pipeline
INPUT_DIR = Path("input")
DATA_DIR = Path("data")
RAW_DUMMY = INPUT_DIR / "raw_dummy.csv"          # <- para no tocar build.py
PARQUET_OUT = DATA_DIR / "summary_allperiods.parquet"

REQUIRED_HINTS = ["Tipo de reporte", "Tipo folio", "Mes", "Periodo"]

def find_header_row_excel(file_bytes: bytes, sheet_name: str = "Base", scan_rows: int = 40) -> int:
    """Encuentra la fila donde están los headers (busca 'Tipo de reporte')."""
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    for r in range(1, scan_rows + 1):
        row_vals = []
        for c in range(1, 80):
            v = ws.cell(r, c).value
            if v is not None:
                row_vals.append(str(v).strip())
        if any(v.lower() == "tipo de reporte" for v in row_vals):
            return r - 1  # pandas usa header 0-based
    return 0

def find_header_row_csv(text: str, scan_lines: int = 60) -> int:
    lines = text.splitlines()
    for i, line in enumerate(lines[:scan_lines]):
        low = line.lower()
        if ("tipo de reporte" in low) and ("tipo folio" in low) and ("mes" in low) and ("periodo" in low):
            return i
    return 0

def normalize_to_raw_dummy(uploaded_file) -> tuple[int, int]:
    """Convierte CSV/XLSX al CSV canónico input/raw_dummy.csv."""
    INPUT_DIR.mkdir(exist_ok=True)
    DATA_DIR.mkdir(exist_ok=True)

    name = (uploaded_file.name or "").lower()
    file_bytes = uploaded_file.getvalue()

    if name.endswith(".xlsx"):
        header_row = find_header_row_excel(file_bytes, sheet_name="Base")
        df = pd.read_excel(BytesIO(file_bytes), sheet_name="Base", header=header_row)
    else:
        text = file_bytes.decode("utf-8-sig", errors="replace")
        header_line = find_header_row_csv(text)
        df = pd.read_csv(BytesIO(file_bytes), encoding="utf-8-sig", skiprows=header_line)

    # Limpieza mínima para evitar mismatches típicos
    if "Tipo folio" in df.columns:
        df["Tipo folio"] = df["Tipo folio"].astype(str).str.strip()
    if "Mes" in df.columns:
        df["Mes"] = df["Mes"].astype(str).str.strip()

    df.to_csv(RAW_DUMMY, index=False)
    return df.shape[0], df.shape[1]


st.title("📤 Cargar base (MVP)")
st.caption("Sube CSV o Excel con el layout del template. Internamente lo convertimos a input/raw_dummy.csv y corremos el pipeline.")

import json
from datetime import datetime
from pathlib import Path

LAST_RUN = Path("data/last_run.json")
PARQUET_OUT = Path("data/summary_allperiods.parquet")  # por si no estaba definido arriba

if LAST_RUN.exists():
    st.info("📌 Última corrida detectada:")
    st.json(json.loads(LAST_RUN.read_text(encoding="utf-8")))

uploaded = st.file_uploader("Sube tu base", type=["csv", "xlsx"])
process = st.button("Procesar base", type="primary", disabled=(uploaded is None))
if process and uploaded is not None:
    # status da feedback claro por etapas
    with st.status("Procesando base…", expanded=True) as status:
        try:
            st.write("1) Normalizando archivo (layout → raw_dummy.csv)…")
            rows, cols = normalize_to_raw_dummy(uploaded)

            st.write("2) Corriendo pipeline (build.py)…")
            runpy.run_path("build.py", run_name="__main__")

            # meta de corrida
            stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            meta = {
                "timestamp": stamp,
                "uploaded_name": getattr(uploaded, "name", None),
                "rows": int(rows),
                "cols": int(cols),
                "parquet_exists": PARQUET_OUT.exists(),
                "parquet_size_mb": round(PARQUET_OUT.stat().st_size / (1024 * 1024), 2) if PARQUET_OUT.exists() else None,
            }

            Path("data").mkdir(exist_ok=True)
            Path("data/last_run.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")

            status.update(label="✅ Pipeline terminado", state="complete", expanded=False)

        except Exception as e:
            status.update(label="❌ Error procesando base", state="error", expanded=True)
            st.exception(e)
            st.stop()

    # Mensaje final CLARO (esto aparece sí o sí al terminar)
    st.success(f"✅ Listo. Base normalizada: {rows:,} filas × {cols:,} columnas.")
    st.switch_page("app.py")   # o "pages/2_Summary.py"


    if PARQUET_OUT.exists():
        st.info("Se generó data/summary_allperiods.parquet.")
        # Opción 1: link
        st.page_link("app.py", label="➡️ Ir al Summary", icon="📊")
        # Si tu summary vive en pages/2_Summary.py, usa esta en vez de app.py:
        # st.page_link("pages/2_Summary.py", label="➡️ Ir al Summary", icon="📊")
    else:
        st.error("No se encontró data/summary_allperiods.parquet. Revisa qué está escribiendo tu build.py.")
